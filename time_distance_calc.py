import googlemaps
import math
from openpyxl import load_workbook
import pgeocode
import xlsxwriter


class TimeDistanceCalculator:

    def __init__(self, API_key, excel_file, worksheet):
        self.API_key = API_key
        self.excel_file = excel_file
        self.wb = load_workbook(excel_file)
        self.ws = self.wb[worksheet]
        self.gmaps = googlemaps.Client(key=API_key)
        self.combined_list = []
        self.orig_chk_combined_list = []
        self.dest_chk_combined_list = []
        self.time_list = []
        self.distance_list = []
        self.userID_list = []

    def dist_time_finder(self, origin, destination):
        '''
        Takes two postcodes and returns the time and distance by road as determined by google maps distance matrix
        '''
        dist_time_matrix = self.gmaps.distance_matrix(origin, destination, mode='driving', units='imperial', region='gb')
        distance = dist_time_matrix['rows'][0]['elements'][0]['distance']['text']
        time = dist_time_matrix['rows'][0]['elements'][0]['duration']['text']
        return time, distance

    def get_data(self, userID_cell, origin_cell, destination_cell):
        '''
        Retrieve data from excel, into lists which are combined and returned as nested lists, each nested list being a row
        '''
        userID_list = [cell.value for cell in self.ws[userID_cell]]
        origin_list = [cell.value for cell in self.ws[origin_cell]]
        destination_list = [cell.value for cell in self.ws[destination_cell]]
        self.combined_list = list(map(list, zip(userID_list, origin_list, destination_list)))
    
    def origin_loc_check(self):
        '''
        Check origin postcode against know area codes for invalid addresses, see if a refined area code is valid and if so append nested list of row data to new checked list
        '''
        nomi = pgeocode.Nominatim('gb')
        for code in self.combined_list:
            if code[1] == None:
                continue
            else:
                checked_postcode = nomi.query_postal_code(code[1][0:4])
                if math.isnan(checked_postcode.loc['latitude']):
                    checked_postcode = nomi.query_postal_code(code[1][0:3])
                    if math.isnan(checked_postcode.loc['latitude']):
                        continue
                    else:
                        code[1] = code[1][0:3]
                        self.orig_chk_combined_list.append(code)
                else:
                    code[1] = code[1][0:4] 
                    self.orig_chk_combined_list.append(code)

    def destination_loc_check(self):
        '''
        Check destination postcode against known area codes for invalid addresses, see if a refined area code is valid and if so append nested list of row data to new checked list
        '''
        nomi = pgeocode.Nominatim('gb')
        for code in self.orig_chk_combined_list:
            if code[2] == None:
                del code
            else:
                checked_postcode = nomi.query_postal_code(code[2])
                if math.isnan(checked_postcode.loc['latitude']):
                    checked_postcode = nomi.query_postal_code(code[2][0:4])
                    if math.isnan(checked_postcode.loc['latitude']):
                        checked_postcode = nomi.query_postal_code(code[2][0:3])
                        if math.isnan(checked_postcode.loc['latitude']):
                            continue
                        else:
                            code[2] = code[2][0:3]
                            self.dest_chk_combined_list.append(code)
                    else:
                        code[2] = code[2][0:4]
                        self.dest_chk_combined_list.append(code)
                else:
                    self.dest_chk_combined_list.append(code)

    def calc_time_dist(self):
        '''
        Put origin and destination postcode into dist_time_finder, then add time, distance and user id to separate lists ready to be written to excel
        '''
        for row in self.dest_chk_combined_list:
            try:
                time, distance = self.dist_time_finder(row[1], row[2])
                self.time_list.append(time)
                self.distance_list.append(distance)
                self.userID_list.append(row[0])
            except(KeyError, TypeError):
                pass 

    def write_to_excel(self, write_file_name):
        '''
        Write user id, travel time and travel distance to rows in a new excel
        '''
        workbook = xlsxwriter.Workbook(write_file_name)
        worksheet = workbook.add_worksheet()
        worksheet.write_column('A1', self.userID_list)
        worksheet.write_column('B1', self.distance_list)
        worksheet.write_column('C1', self.time_list)
        workbook.close()

    def postcode_running(self, userID_cell, origin_cell, destination_cell, write_file_name):
        self.get_data(userID_cell, origin_cell, destination_cell)
        self.origin_loc_check()
        self.destination_loc_check()
        print(self.dest_chk_combined_list)
        self.calc_time_dist()
        self.write_to_excel(write_file_name)
        